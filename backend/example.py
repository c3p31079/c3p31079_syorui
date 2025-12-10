from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# -------------------------
# パス設定
# -------------------------
BASE_DIR = os.path.dirname(__file__)
TEMPLATE_PATH = os.path.join(BASE_DIR, "template.xlsx")
ICON_DIR = os.path.join(BASE_DIR, "icons")
OUTPUT_PATH = os.path.join(BASE_DIR, "output.xlsx")

# -------------------------
# Excelテンプレート作成
# -------------------------
def create_excel_template():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws.title = "点検チェックシート"
    return wb, ws

# -------------------------
# アイテム適用
# -------------------------
def apply_items(ws, items):
    for item in items:
        if "cell" not in item or "type" not in item:
            continue

        cell = item["cell"]

        # -----------------
        # テキスト
        # -----------------
        if item["type"] == "text":
            ws[cell].value = str(item.get("text", ""))
            continue

        # -----------------
        # PNGアイコン
        # -----------------
        if item["type"] == "icon":
            icon_file = item.get("icon")
            if not icon_file:
                continue

            icon_path = os.path.join(ICON_DIR, icon_file)
            if not os.path.exists(icon_path):
                print(f"[WARN] icon not found: {icon_path}")
                continue

            _insert_image(ws, cell, icon_path)

# -------------------------
# 画像挿入関数（openpyxl 3.1.5対応）
# -------------------------
def _insert_image(ws, cell, icon_path):
    img = Image(icon_path)
    # openpyxl 3.1.5 では anchor にセル指定するだけでOK
    img.anchor = cell
    ws.add_image(img)

# -------------------------
# 実行例
# -------------------------
if __name__ == "__main__":
    wb, ws = create_excel_template()

    # サンプルデータ
    items = [
        {"cell": "B2", "type": "text", "text": "点検項目1"},
        {"cell": "C3", "type": "icon", "icon": "check.png"},  # icons/check.png を用意
        {"cell": "D5", "type": "text", "text": "点検項目2"},
        {"cell": "E6", "type": "icon", "icon": "warn.png"}    # icons/warn.png を用意
    ]

    apply_items(ws, items)

    # 保存
    wb.save(OUTPUT_PATH)
    print(f"Excel saved: {OUTPUT_PATH}")
