# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
# from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
# from openpyxl.utils import column_index_from_string
# import os

# BASE_DIR = os.path.dirname(__file__)
# TEMPLATE_PATH = os.path.join(BASE_DIR, "template.xlsx")
# ICON_DIR = os.path.join(BASE_DIR, "icons")

# EMU = 9525          # 1px = 9525 EMU
# ICON_PX = 32        # PNGサイズ（32x32）

# def create_excel_template():
#     wb = load_workbook(TEMPLATE_PATH)
#     ws = wb.active
#     ws.title = "点検チェックシート"
#     return wb, ws

# def apply_items(ws, items):
#     for item in items:
#         if "cell" not in item or "type" not in item:
#             continue

#         cell = item["cell"]

#         # ---------- テキスト ----------
#         if item["type"] == "text":
#             ws[cell].value = str(item.get("text", ""))
#             continue

#         # ---------- アイコン ----------
#         if item["type"] == "icon":
#             icon_file = item.get("icon")
#             if not icon_file:
#                 continue

#             icon_path = os.path.join(ICON_DIR, icon_file)
#             if not os.path.exists(icon_path):
#                 print(f"[WARN] icon not found: {icon_path}")
#                 continue

#             _insert_image(
#                 ws,
#                 cell,
#                 icon_path,
#                 dx=item.get("dx", 0),
#                 dy=item.get("dy", 0)
#             )

# def _insert_image(ws, cell, icon_path, dx=0, dy=0):
#     """
#     32x32 PNG を dx,dy(px) 指定でセル内に配置
#     """
#     img = Image(icon_path)

#     # 明示的にサイズ指定（重要）
#     img.width = ICON_PX
#     img.height = ICON_PX

#     col_letter = ''.join(filter(str.isalpha, cell))
#     row_number = int(''.join(filter(str.isdigit, cell)))

#     col = column_index_from_string(col_letter) - 1
#     row = row_number - 1

#     marker = AnchorMarker(
#         col=col,
#         colOff=dx * EMU,
#         row=row,
#         rowOff=dy * EMU
#     )

#     # ★ ext を必ず指定する ★
#     anchor = OneCellAnchor(
#         _from=marker,
#         ext=(ICON_PX * EMU, ICON_PX * EMU)
#     )

#     img.anchor = anchor
#     ws.add_image(img)

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import column_index_from_string
import os

BASE_DIR = os.path.dirname(__file__)
TEMPLATE_PATH = os.path.join(BASE_DIR, "template.xlsx")
ICON_DIR = os.path.join(BASE_DIR, "icons")

EMU = 9525
ICON_PX = 32

def create_excel_template():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws.title = "点検チェックシート"
    return wb, ws

def apply_items(ws, items):
    for item in items:
        if "cell" not in item or "type" not in item:
            continue

        cell = item["cell"]

        # ---------- テキスト ----------
        if item["type"] == "text":
            target_cell = _get_top_left_cell(ws, cell)
            target_cell.value = str(item.get("text", ""))
            continue

        # ---------- アイコン ----------
        if item["type"] == "icon":
            icon_file = item.get("icon")
            if not icon_file:
                continue
            icon_path = os.path.join(ICON_DIR, icon_file)
            if not os.path.exists(icon_path):
                print(f"[WARN] icon not found: {icon_path}")
                continue
            _insert_image(ws, cell, icon_path, dx=item.get("dx", 0), dy=item.get("dy", 0))

def _get_top_left_cell(ws, cell_str):
    """
    マージセルの場合は左上セルを返す
    """
    for merged in ws.merged_cells.ranges:
        if cell_str in str(merged):
            return ws.cell(row=merged.min_row, column=merged.min_col)
    return ws[cell_str]

def _insert_image(ws, cell, icon_path, dx=0, dy=0):
    img = Image(icon_path)
    img.width = ICON_PX
    img.height = ICON_PX

    col_letter = ''.join(filter(str.isalpha, cell))
    row_number = int(''.join(filter(str.isdigit, cell)))

    col = column_index_from_string(col_letter) - 1
    row = row_number - 1

    marker = AnchorMarker(col=col, colOff=dx * EMU, row=row, rowOff=dy * EMU)
    anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(ICON_PX * EMU, ICON_PX * EMU))
    img.anchor = anchor
    ws.add_image(img)
