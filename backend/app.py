from flask import Flask, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
from flask_cors import CORS
import io
import os

EMU = 9525
ICON_PX = 16

BASE_DIR = os.path.dirname(__file__)
TEMPLATE_PATH = os.path.join(BASE_DIR, "template.xlsx")
ICON_DIR = os.path.join(BASE_DIR, "icons")

app = Flask(__name__)
CORS(app)


def insert_icon(ws, cell, icon_file, dx=0, dy=0):
    img_path = os.path.join(ICON_DIR, icon_file)
    if not os.path.exists(img_path): return
    img = Image(img_path)
    img.width = ICON_PX
    img.height = ICON_PX

    col_letter = ''.join(filter(str.isalpha, cell))
    row_number = int(''.join(filter(str.isdigit, cell)))
    col_idx = column_index_from_string(col_letter) - 1

    marker = AnchorMarker(col=col_idx, colOff=dx*EMU,
                          row=row_number-1, rowOff=dy*EMU)
    anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(EMU*img.width, EMU*img.height))
    anchor.graphicFrame = img._data
    ws.add_image(img, cell)


def insert_text(ws, cell, text):
    col_letter = ''.join(filter(str.isalpha, cell))
    row_number = int(''.join(filter(str.isdigit, cell)))
    col = column_index_from_string(col_letter) - 1
    row = row_number - 1

    # マージセル対応
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            cell = merged_range.start_cell.coordinate
            col_letter = ''.join(filter(str.isalpha, cell))
            row_number = int(''.join(filter(str.isdigit, cell)))
            col = column_index_from_string(col_letter) - 1
            row = row_number - 1
            break

    ws.cell(row=row + 1, column=col + 1, value=text)
    ws.cell(row=row + 1, column=col + 1).alignment = Alignment(wrap_text=True, vertical='top')


@app.route("/api/generate_excel", methods=["POST"])
def generate_excel():
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({"error": "JSONが正しく送信されていません"}), 400

    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({"error": "テンプレートファイルが見つかりません"}), 500

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    for item in data.get("items", []):
        cell = item.get("cell")
        if not cell:
            continue

        item_type = item.get("type")
        dx = item.get("dx", 0)
        dy = item.get("dy", 0)

        if item_type == "icon" and item.get("icon"):
            insert_icon(ws, cell, item["icon"], dx=dx, dy=dy)
        elif item_type in ("text", "number") and item.get("value") is not None:
            insert_text(ws, cell, str(item["value"]))
        elif item_type == "checkbox":
            if item.get("value"):
                insert_icon(ws, cell, item.get("icon", "check.png"), dx=dx, dy=dy)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="点検チェックシート.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(debug=True)
